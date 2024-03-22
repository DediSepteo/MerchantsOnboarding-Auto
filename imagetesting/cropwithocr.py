'''
Scanning of menu by cropping and storing in excel file
'''

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from PIL import ImageTk, Image, ImageEnhance
import easyocr
import numpy as np
import openpyxl
import os


class ImageUploaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Image Uploader")

        self.canvas = tk.Canvas(self.root, width=1200, height=900)
        self.canvas.pack()

        self.upload_button = tk.Button(self.root, text="Upload Image", command=self.upload_image)
        self.upload_button.pack()

        self.ocr_button = tk.Button(self.root, text="Perform OCR", command=self.perform_ocr)
        self.ocr_button.pack()

        self.reader = easyocr.Reader(['en', 'ch_sim'])  # Initialize EasyOCR reader for English language
        self.start_x = None
        self.start_y = None
        self.rectangle = None
        self.cropped_image = None

        self.excel_file = None  # Store the Excel file path

        # Ask the user whether to open an existing Excel file or create a new one
        self.open_or_create_excel()

    def open_or_create_excel(self):
        choice = messagebox.askyesno("Excel File", "Do you want to open an existing Excel file?")
        if choice:
            self.open_existing_excel()
        else:
            self.create_new_excel()

    def open_existing_excel(self):
        file_path = filedialog.askopenfilename(title="Open Existing Excel File",
                                            filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.excel_file = file_path


    def create_new_excel(self):
        file_path = filedialog.asksaveasfilename(title="Save As",
                                                 defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.excel_file = file_path

    # Other methods remain unchanged


    def upload_image(self):
        file_path = filedialog.askopenfilename(title="Select Image",
                                               filetypes=(("Image files", "*.png;*.jpg;*.jpeg;*.gif"), ("All files", "*.*")))
        if file_path:
            self.image = Image.open(file_path)
            self.photo = ImageTk.PhotoImage(self.image)
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
            self.canvas.bind("<ButtonPress-1>", self.start_cropping)
            self.canvas.bind("<B1-Motion>", self.draw_rectangle)
            self.canvas.bind("<ButtonRelease-1>", self.end_cropping)

    def start_cropping(self, event):
        self.start_x = self.canvas.canvasx(event.x)
        self.start_y = self.canvas.canvasy(event.y)
        self.rectangle = self.canvas.create_rectangle(self.start_x, self.start_y, self.start_x, self.start_y, outline="red")

    def draw_rectangle(self, event):
        cur_x = self.canvas.canvasx(event.x)
        cur_y = self.canvas.canvasy(event.y)
        self.canvas.coords(self.rectangle, self.start_x, self.start_y, cur_x, cur_y)

    def end_cropping(self, event):
        end_x = self.canvas.canvasx(event.x)
        end_y = self.canvas.canvasy(event.y)
        
        # Calculate the width and height of the cropped area
        width = end_x - self.start_x
        height = end_y - self.start_y
        
        # Determine the target width and height while maintaining the aspect ratio
        target_width = 980
        target_height = int(height * (target_width / width))

        # Check if cropped_image exists
        if self.image:
            # Crop the image
            self.cropped_image = self.image.crop((self.start_x, self.start_y, end_x, end_y))

            # Resize the cropped image while preserving its aspect ratio
            self.cropped_image = self.cropped_image.resize((target_width, target_height))
            self.cropped_image = self.sharpen_image(self.cropped_image)  # Sharpen the cropped image

            self.cropped_photo = ImageTk.PhotoImage(self.cropped_image)
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.cropped_photo)

    def perform_ocr(self):
        if self.cropped_image:
            enhanced_image = ImageEnhance.Contrast(self.cropped_image).enhance(1.5)

            # Convert the cropped image to a numpy array
            cropped_image_array = np.array(enhanced_image)

            # Perform OCR on the numpy array
            result = self.reader.readtext(cropped_image_array, detail=0)

            if self.excel_file:
                # Open existing or create new workbook
                wb = openpyxl.load_workbook(self.excel_file) if os.path.exists(self.excel_file) else openpyxl.Workbook()
                ws = wb.active
                ws.title = "OCR Results"

                # Write OCR results to Excel
                for i, text in enumerate(result, start=1):
                    ws.cell(row=i, column=1, value=text)

                # Save the workbook
                wb.save(self.excel_file)
                print("OCR results saved to", self.excel_file)
            else:
                print("No Excel file selected.")
        else:
            print("No cropped image available for OCR")

    
def main():
    root = tk.Tk()
    app = ImageUploaderApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
